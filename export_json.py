from openpyxl import load_workbook
from datetime import datetime, date
import json
from pathlib import Path
import os

# Force everything to use your Desktop
desktop = Path.home() / "Desktop"
excel_file = desktop / "Test -Overseas Leave Application.xlsx"
output_file = desktop / "data.json"
sheet_name = "Daily Slots"

def normalize_date(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if value is None:
        return ""
    return str(value).strip()

def main():
    print("Current working folder:", os.getcwd())
    print("Looking for Excel file at:", excel_file)
    print("Will save JSON to:", output_file)

    if not excel_file.exists():
      print("ERROR: Excel file not found.")
      return

    wb = load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]

    result = []
    row = 2

    while True:
        date_val = ws[f"A{row}"].value
        quota_bal = ws[f"C{row}"].value

        if date_val in (None, ""):
            break

        result.append({
            "date": normalize_date(date_val),
            "quotaBalance": quota_bal
        })
        row += 1

    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"SUCCESS: Exported {len(result)} rows to {output_file}")

if __name__ == "__main__":
    main()